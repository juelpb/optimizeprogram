""" 
Functions for generating the parametrized girder section geometry. 
"""

import numpy as np
from matplotlib import pyplot as plt
import pandas as pd
from openpyxl import Workbook


def GenerateCorners(H, p_type):
    """Function to generate the corner coordinates along the mid-line of the parametrized girder section

    Args:
        H (float): Girder height
        p_type (int): Parametrization type. Girder sections from earlier Langenuen studies are implemented. 2020, 2021, 2022

    Returns:
        x_list, y_list: List of corner coordinates
    """
    if p_type == 2021:
        A = 9.538
        B = 30.64
        C = 1
        D = 4.451
        theta = (A*H-B)/(C*H-D)
        theta = theta*2*np.pi/360
    
    elif p_type == 2020:
        A = 2.1404
        B = 7.5532
        C = 0.2062
        D = 1
        theta = (A*H-B)/(C*H-D)
        theta = theta*2*np.pi/360

    if p_type == 2022:
        theta = 16 # In degrees
        theta = theta*2*np.pi/360


    if p_type == 2021 or p_type == 2020:
        # Points along mid-line
        x1 = 0
        y1 = 0

        x2 = 31/2
        y2 = -0.03*31/2

        x3 = x2+0.4*H
        y3 = y2-np.tan(np.pi/6)*0.4*H

        x4 = x3-(H+y3)/np.tan(theta)
        y4 = -H
        
        x7 = -31/2
        y7 = -0.03*31/2

        x6 = x7-0.4*H
        y6 = y7-np.tan(np.pi/6)*0.4*H

        x5 = x6+(H+y6)/np.tan(theta)
        y5 = -H

        x8 = x1
        y8 = y1
        
    elif p_type == 2022:
        h_incU2 = 1.310
        h_incU1 = 1.190
        h_fall2 = 0.330
        h_fall1 = 0.450
        h_incL = H - h_fall1 - h_incU1

        w_t = 31
        w_l2 = 11
        w_l1 = 15
        w_incU = (w_t-w_l1-w_l2)/2
        w_incL = h_incL/np.tan(theta)
        w_L = w_t - 2*w_incL + 2*w_incU

        # Coordinates
        x1 = 0
        y1 = 0

        x2 = w_l1
        y2 = -h_fall1

        x3 = x2 + w_incU
        y3 = y2 - h_incU1

        x4 = x3 - w_incL
        y4 = -H

        x7 = -w_l2
        y7 = -h_fall2

        x6 = x7 - w_incU
        y6 = y7 - h_incU2

        x5 = x6 + w_incL
        y5 = -H

        x8 = x1
        y8 = y1
        
    xs = [x6, x7, x1, x2, x3, x4, x5, x6]
    ys = [y6, y7, y1, y2, y3, y4, y5, y6]
    
    return xs, ys

# Offsetting functions
#____________________________________________________________________
def normalizeVec(x,y):
    """Generating normalvectors to the input vectors

    Args:
        x (list or array): x-coordinates of input vectors
        y (list or array): y-coordinates of input vectors

    Returns:
        list or array: x-coordinates and y-coordinates of normalvectors
    """
    normx = x/np.sqrt(x**2+y**2+1e-10)
    normy = y/np.sqrt(x**2+y**2+1e-10)
    
    return normx, normy


def makeOffsetPoly(oldX, oldY, offset, outer_ccw = 1):
    """_summary_

    Args:
        oldX (list): x-coordinates of lines to be offset
        oldY (list): y-coordinates of lines to be offset
        offset (float): offset
        outer_ccw (int, optional): _description_. Defaults to 1.

    Returns:
        _type_: _description_
    """

    num_points = len(oldX)
    newX = []
    newY = []
    
    for curr in range(num_points):
        prev = (curr + num_points - 1) % num_points
        next = (curr + 1) % num_points

        vnX =  oldX[next] - oldX[curr]
        vnY =  oldY[next] - oldY[curr]
        vnnX, vnnY = normalizeVec(vnX,vnY)
        nnnX = vnnY
        nnnY = -vnnX

        vpX =  oldX[curr] - oldX[prev]
        vpY =  oldY[curr] - oldY[prev]
        vpnX, vpnY = normalizeVec(vpX,vpY)
        npnX = vpnY * outer_ccw
        npnY = -vpnX * outer_ccw

        bisX = (nnnX + npnX) * outer_ccw
        bisY = (nnnY + npnY) * outer_ccw

        bisnX, bisnY = normalizeVec(bisX,  bisY)
        bislen = offset /  np.sqrt(1 + nnnX*npnX + nnnY*npnY)

        newX.append(oldX[curr] + bislen * bisnX)
        newY.append(oldY[curr] + bislen * bisnY)

    return newX, newY

"""
# Initial parameters
H = 6.1
tef = 0.0445
p_type = 2021   # Parametrization type; 2020, 2021 or 2022



#___________________________________________________________________________
# Plotting example

off = tef/2

x = GenerateCorners(H, p_type)[0]
y = GenerateCorners(H, p_type)[1]

x_i, y_i = makeOffsetPoly(x, y, -off, outer_ccw = -1)
x_o, y_o = makeOffsetPoly(x, y, off)

temp_x = (x_i[0]+x_i[-1])/2
x_i[0] = temp_x
x_i[-1] = temp_x

temp_x = (x_o[0]+x_o[-1])/2
x_o[0] = temp_x
x_o[-1] = temp_x

temp_y = (y_i[0]+y_i[-1])/2
y_i[0] = temp_y
y_i[-1] = temp_y

temp_y = (y_o[0]+y_o[-1])/2
y_o[0] = temp_y
y_o[-1] = temp_y


pts_i = np.zeros([len(x_i),2])
pts_o = np.zeros([len(x_i),2])
for i in range(len(x_i)):
    pts_i[i,0] = x_i[i]
    pts_i[i,1] = y_i[i]
    
for i in range(len(x_o)):
    pts_o[i,0] = x_o[i]
    pts_o[i,1] = y_o[i]
  
plt.plot(x_i, y_i)  
plt.plot(x_o, y_o)
plt.ylim([-29, 1])
plt.show()
"""



#______________________________________________________________________________
# Writing coordinates to excel

"""
coords = np.zeros([5, 7, 2])

# Full scale
hs = [3.5, 3.75, 4.0, 4.25, 4.5]
for sec in range(len(hs)):
    xs = GenerateCorners(hs[sec], 2022)[0]
    ys = GenerateCorners(hs[sec], 2022)[1]
    for i in range(len(xs)-1):
        coords[sec,i,0] = xs[i]
        coords[sec,i,1] = ys[i]

wb = Workbook()

path = 'C:/Users/sverr/OneDrive - NTNU/My files/Langenuen_masteroppgave/Tverrsnitt/Tverrsnitt_22_koord.xls'

wb.active
for sec in range(len(hs)):
    data = wb.create_sheet(title=f'h={hs[sec]}')
    data.cell(column=1, row=1, value='Corner number')
    data.cell(column=2, row=1, value='x')
    data.cell(column=3, row=1, value='y')

    for i in range(len(coords[0])):
        data.cell(column=1, row=i+2, value=i+1)

    for i in range(len(coords[0])):
        for j in range(len(coords[0][0])):
            data.cell(column=j+2, row=i+2, value=coords[sec,i,j])
        
wb.save(filename=path)



# Model scale
for sec in range(len(hs)):
    xs = GenerateCorners(hs[sec], 2022)[0]
    ys = GenerateCorners(hs[sec], 2022)[1]
    for i in range(len(xs)-1):
        coords[sec,i,0] = xs[i]/70
        coords[sec,i,1] = ys[i]/70

wb = Workbook()

path = 'C:/Users/sverr/OneDrive - NTNU/My files/Langenuen_masteroppgave/Tverrsnitt/Tverrsnitt_22_koord_modscale.xls'

wb.active
for sec in range(len(hs)):
    data = wb.create_sheet(title=f'h={hs[sec]}')
    data.cell(column=1, row=1, value='Corner number')
    data.cell(column=2, row=1, value='x')
    data.cell(column=3, row=1, value='y')

    for i in range(len(coords[0])):
        data.cell(column=1, row=i+2, value=i+1)

    for i in range(len(coords[0])):
        for j in range(len(coords[0][0])):
            data.cell(column=j+2, row=i+2, value=coords[sec,i,j])
        
wb.save(filename=path)
"""